from __future__ import annotations

from datetime import datetime
from io import BytesIO
import re
from typing import Any

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from api.services.dashboard_service import DashboardService


class DashboardReportService:
    _QUARTER_PATTERN = re.compile(r"^(20\d{2})-Q([1-4])$")
    _HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    _SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
    _SUBTLE_FILL = PatternFill("solid", fgColor="F8FAFC")
    _HEADER_FONT = Font(color="FFFFFF", bold=True)
    _SECTION_FONT = Font(bold=True, color="1F2937")
    _BOLD_FONT = Font(bold=True)
    _CENTER = Alignment(horizontal="center", vertical="center")
    _LEFT = Alignment(horizontal="left", vertical="center")
    _THIN_BORDER = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    _MONEY_FORMAT = '$#,##0.00;[Red]-$#,##0.00;-'
    _PCT_FORMAT = '0.0%;[Red]-0.0%;-'

    def __init__(self, dashboard_service: DashboardService | None = None):
        self.dashboard_service = dashboard_service or DashboardService()

    def build_report(
        self,
        *,
        period_type: str,
        period: str,
        limit: int = 10,
        include_details: bool = True,
        db: Session,
    ) -> tuple[BytesIO, str]:
        normalized_period_type = str(period_type or "").strip().lower()
        self._validate_period(normalized_period_type, period)

        summary_payload = self.dashboard_service.get_main_stats(db=db)
        insights_payload = self._normalize_insights(self.dashboard_service.get_insights(db=db))
        period_summary = self._build_period_summary(
            period_type=normalized_period_type,
            period=period,
            trend=summary_payload.get("trend", []),
            db=db,
        )
        top_clients_payload = self._get_period_clients(
            period_type=normalized_period_type,
            period=period,
            limit=limit,
            db=db,
        )
        details_payload = None
        if include_details:
            details_payload = self._get_period_clients(
                period_type=normalized_period_type,
                period=period,
                limit=100000,
                db=db,
            )

        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        self._write_summary_sheet(
            sheet=summary_sheet,
            period_type=normalized_period_type,
            period=period,
            summary_payload=summary_payload,
            period_summary=period_summary,
            insights_payload=insights_payload,
            include_details=include_details,
        )
        self._write_clients_sheet(
            sheet=workbook.create_sheet("Top Clients"),
            title="Top Clients",
            subtitle=f"{period} TOP {limit} 客户对比",
            compare_label=period_summary["compare_label"],
            clients=top_clients_payload.get("clients", []),
        )

        if include_details:
            self._write_clients_sheet(
                sheet=workbook.create_sheet("Details"),
                title="Details",
                subtitle=f"{period} 全量客户明细",
                compare_label=period_summary["compare_label"],
                clients=(details_payload or {}).get("clients", []),
            )

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer, f"dashboard_report_{period}.xlsx"

    def _validate_period(self, period_type: str, period: str) -> None:
        if period_type not in {"month", "quarter"}:
            raise HTTPException(status_code=400, detail="period_type must be 'month' or 'quarter'")
        if period_type == "month" and not self.dashboard_service._is_valid_month(period):
            raise HTTPException(status_code=400, detail="Invalid month period. Expected YYYY-MM")
        if period_type == "quarter" and not self._QUARTER_PATTERN.match(str(period or "").strip()):
            raise HTTPException(status_code=400, detail="Invalid quarter period. Expected YYYY-QN")

    def _normalize_insights(self, payload: dict[str, Any] | None) -> dict[str, Any]:
        metrics = (payload or {}).get("metrics") or {}
        segmentation = (payload or {}).get("segmentation") or {}
        return {
            "metrics": {
                "anomalies": metrics.get("anomalies", []),
                "churn_risk": metrics.get("churn_risk", []),
                "growth_stars": metrics.get("growth_stars", []),
            },
            "segmentation": {
                "whales": segmentation.get("whales", {"count": 0, "value": 0.0, "pct": 0.0}),
                "core": segmentation.get("core", {"count": 0, "value": 0.0, "pct": 0.0}),
                "long_tail": segmentation.get("long_tail", {"count": 0, "value": 0.0, "pct": 0.0}),
            },
        }

    def _get_period_clients(self, *, period_type: str, period: str, limit: int, db: Session) -> dict[str, Any]:
        if period_type == "month":
            return self.dashboard_service.get_month_top_clients(period, limit, db, compare_prev=True)
        return self.dashboard_service.get_quarter_top_clients(period, limit, db, compare_prev=True)

    def _build_period_summary(
        self,
        *,
        period_type: str,
        period: str,
        trend: list[dict[str, Any]],
        db: Session,
    ) -> dict[str, Any]:
        month_totals = {
            item["month"]: {
                "consumption": float(item.get("total_consumption") or 0.0),
                "fee": float(item.get("total_service_fee") or 0.0),
            }
            for item in trend
            if item.get("month")
        }

        if period_type == "month":
            compare_label = self.dashboard_service._get_previous_month(period)
            if compare_label and compare_label not in month_totals:
                compare_label = self.dashboard_service._find_previous_available_month(period, db)
            current = month_totals.get(period, {"consumption": 0.0, "fee": 0.0})
            compare = month_totals.get(compare_label or "", {"consumption": 0.0, "fee": 0.0})
            return {
                "period_label": period,
                "compare_label": compare_label or "无可用对比周期",
                "months": [period],
                "current_consumption": current["consumption"],
                "current_fee": current["fee"],
                "compare_consumption": compare["consumption"] if compare_label else 0.0,
                "compare_fee": compare["fee"] if compare_label else 0.0,
            }

        quarter_months_map = self.dashboard_service._get_quarter_months_map(db)
        months = quarter_months_map.get(period, [])
        compare_label = self.dashboard_service._get_previous_quarter(period)
        if compare_label and not quarter_months_map.get(compare_label):
            earlier = [key for key in sorted(quarter_months_map.keys()) if key < period]
            compare_label = earlier[-1] if earlier else None
        compare_months = quarter_months_map.get(compare_label or "", [])

        current_consumption, current_fee = self._sum_period(months, month_totals)
        compare_consumption, compare_fee = self._sum_period(compare_months, month_totals)
        return {
            "period_label": period,
            "compare_label": compare_label or "无可用对比周期",
            "months": months,
            "current_consumption": current_consumption,
            "current_fee": current_fee,
            "compare_consumption": compare_consumption,
            "compare_fee": compare_fee,
        }

    def _sum_period(self, months: list[str], month_totals: dict[str, dict[str, float]]) -> tuple[float, float]:
        consumption = 0.0
        fee = 0.0
        for month in months:
            values = month_totals.get(month)
            if not values:
                continue
            consumption += values["consumption"]
            fee += values["fee"]
        return consumption, fee

    def _write_summary_sheet(
        self,
        *,
        sheet,
        period_type: str,
        period: str,
        summary_payload: dict[str, Any],
        period_summary: dict[str, Any],
        insights_payload: dict[str, Any],
        include_details: bool,
    ) -> None:
        stats = summary_payload.get("stats") or {}
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = "对外汇报看板报表"
        sheet["A1"].font = Font(size=16, bold=True, color="1F2937")
        sheet["A2"] = "生成时间"
        sheet["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet["A3"] = "周期类型"
        sheet["B3"] = "月度" if period_type == "month" else "季度"
        sheet["A4"] = "导出周期"
        sheet["B4"] = period
        sheet["A5"] = "包含明细"
        sheet["B5"] = "是" if include_details else "否"
        sheet["A6"] = "数据来源"
        sheet["B6"] = "/api/dashboard | /api/dashboard/insights | /api/dashboard/month/{month}/top-clients | /api/dashboard/quarter/{quarter}/top-clients"
        for cell in ("A2", "A3", "A4", "A5", "A6"):
            sheet[cell].font = self._BOLD_FONT

        row = 8
        row = self._write_section_title(sheet, row, "KPI 总览")
        self._write_table_headers(sheet, row, ["指标", "值", "说明"])
        row += 1

        kpi_rows = [
            ("本期总消耗", stats.get("consumption"), f"最新月份：{stats.get('month', '-') }"),
            ("本期服务费", stats.get("fee"), "来自 /api/dashboard 主看板统计"),
            ("消耗环比", self._pct_to_decimal(stats.get("consumption_mom")), "最新月份对比上一可用月份"),
            ("服务费环比", self._pct_to_decimal(stats.get("fee_mom")), "最新月份对比上一可用月份"),
            ("消耗同比", self._pct_to_decimal(stats.get("consumption_yoy")), "最新月份对比去年同月"),
            ("服务费同比", self._pct_to_decimal(stats.get("fee_yoy")), "最新月份对比去年同月"),
        ]
        for name, value, note in kpi_rows:
            sheet.cell(row=row, column=1, value=name)
            value_cell = sheet.cell(row=row, column=2, value=value if value is not None else "-")
            sheet.cell(row=row, column=3, value=note)
            self._apply_body_style(sheet, row, 3)
            if "环比" in name or "同比" in name:
                value_cell.number_format = self._PCT_FORMAT
            elif isinstance(value, (int, float)):
                value_cell.number_format = self._MONEY_FORMAT
            row += 1

        row += 1
        row = self._write_section_title(sheet, row, "当前周期摘要")
        self._write_table_headers(sheet, row, ["项目", "当前周期", "对比周期", "差值"])
        row += 1

        current_consumption = float(period_summary["current_consumption"])
        current_fee = float(period_summary["current_fee"])
        compare_consumption = float(period_summary["compare_consumption"])
        compare_fee = float(period_summary["compare_fee"])
        summary_rows = [
            ("周期标签", period_summary["period_label"], period_summary["compare_label"], "—"),
            ("覆盖月份", ", ".join(period_summary["months"]) if period_summary["months"] else "无数据", "—", "—"),
            ("总消耗", current_consumption, compare_consumption, current_consumption - compare_consumption),
            ("总服务费", current_fee, compare_fee, current_fee - compare_fee),
        ]
        for name, current_value, compare_value, delta in summary_rows:
            sheet.cell(row=row, column=1, value=name)
            current_cell = sheet.cell(row=row, column=2, value=current_value)
            compare_cell = sheet.cell(row=row, column=3, value=compare_value)
            delta_cell = sheet.cell(row=row, column=4, value=delta)
            self._apply_body_style(sheet, row, 4)
            if name in {"总消耗", "总服务费"}:
                current_cell.number_format = self._MONEY_FORMAT
                compare_cell.number_format = self._MONEY_FORMAT
                delta_cell.number_format = self._MONEY_FORMAT
            row += 1

        row += 1
        row = self._write_section_title(sheet, row, "洞察摘要")
        self._write_table_headers(sheet, row, ["分类", "客户", "说明", "数值"])
        row += 1

        insight_rows = self._flatten_insights(insights_payload)
        if not insight_rows:
            sheet.cell(row=row, column=1, value="系统提示")
            sheet.cell(row=row, column=2, value="—")
            sheet.cell(row=row, column=3, value="当前暂无可导出的洞察数据")
            sheet.cell(row=row, column=4, value="—")
            self._apply_body_style(sheet, row, 4)
            row += 1
        else:
            for category, client, detail, metric in insight_rows:
                sheet.cell(row=row, column=1, value=category)
                sheet.cell(row=row, column=2, value=client)
                sheet.cell(row=row, column=3, value=detail)
                metric_cell = sheet.cell(row=row, column=4, value=metric)
                self._apply_body_style(sheet, row, 4)
                if isinstance(metric, (int, float)):
                    metric_cell.number_format = self._MONEY_FORMAT
                row += 1

        self._set_widths(sheet, {1: 18, 2: 20, 3: 42, 4: 18})

    def _flatten_insights(self, payload: dict[str, Any]) -> list[tuple[str, str, str, Any]]:
        rows: list[tuple[str, str, str, Any]] = []
        metrics = payload["metrics"]
        segmentation = payload["segmentation"]

        for item in metrics["anomalies"]:
            direction = "暴增" if item.get("type") == "surge" else "暴跌"
            rows.append(("异动预警", item.get("client", "-"), f"{direction} {abs(float(item.get('change_pct') or 0.0)):.1f}%", item.get("value", 0.0)))
        for item in metrics["churn_risk"]:
            if item.get("trend") == "declining":
                detail = f"连续 {item.get('consecutive_months', 0)} 月下滑"
            else:
                detail = "本月无消耗"
            rows.append(("流失风险", item.get("client", "-"), detail, item.get("value", 0.0)))
        for item in metrics["growth_stars"]:
            rows.append(("增长之星", item.get("client", "-"), "增长潜力客户", item.get("growth_amount", 0.0)))

        for key, label in (("whales", "头部客户"), ("core", "腰部客户"), ("long_tail", "长尾客户")):
            segment = segmentation.get(key, {})
            rows.append((
                "客户分层",
                label,
                f"客户数 {int(segment.get('count', 0))}，占比 {float(segment.get('pct', 0.0)):.1f}%",
                float(segment.get("value", 0.0)),
            ))
        return rows

    def _write_clients_sheet(
        self,
        *,
        sheet,
        title: str,
        subtitle: str,
        compare_label: str,
        clients: list[dict[str, Any]],
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = title
        sheet["A1"].font = Font(size=16, bold=True, color="1F2937")
        sheet["A2"] = subtitle
        sheet["A3"] = f"对比周期：{compare_label}"

        header_row = 5
        headers = ["排名", "客户", "本期消耗", "本期服务费", "上期消耗", "上期服务费", "消耗增量", "服务费增量", "排名变化"]
        self._write_table_headers(sheet, header_row, headers)
        row = header_row + 1

        if not clients:
            sheet.cell(row=row, column=1, value="暂无数据")
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            merged = sheet.cell(row=row, column=1)
            merged.alignment = self._CENTER
            merged.font = Font(color="6B7280", italic=True)
            merged.fill = self._SUBTLE_FILL
            merged.border = self._THIN_BORDER
            row += 1
        else:
            for idx, item in enumerate(clients, start=1):
                values = [
                    item.get("rank", idx),
                    item.get("client_name", ""),
                    float(item.get("consumption") or 0.0),
                    float(item.get("service_fee") or 0.0),
                    item.get("prev_consumption"),
                    item.get("prev_service_fee"),
                    item.get("consumption_delta"),
                    item.get("fee_delta"),
                    item.get("rank_change"),
                ]
                for column, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row, column=column, value=value)
                    cell.border = self._THIN_BORDER
                    cell.alignment = self._LEFT if column == 2 else self._CENTER
                    if column >= 3 and column <= 8 and isinstance(value, (int, float)):
                        cell.number_format = self._MONEY_FORMAT
                    if column == 9 and isinstance(value, (int, float)):
                        cell.number_format = '0;[Red]-0;0'
                row += 1

        sheet.freeze_panes = "A6"
        sheet.auto_filter.ref = f"A5:I{max(row - 1, 5)}"
        self._set_widths(sheet, {1: 8, 2: 24, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 10})

    def _write_section_title(self, sheet, row: int, title: str) -> int:
        sheet.cell(row=row, column=1, value=title)
        sheet.cell(row=row, column=1).font = self._SECTION_FONT
        sheet.cell(row=row, column=1).fill = self._SECTION_FILL
        sheet.cell(row=row, column=1).alignment = self._LEFT
        sheet.cell(row=row, column=1).border = self._THIN_BORDER
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        return row + 1

    def _write_table_headers(self, sheet, row: int, headers: list[str]) -> None:
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column, value=header)
            cell.fill = self._HEADER_FILL
            cell.font = self._HEADER_FONT
            cell.alignment = self._CENTER
            cell.border = self._THIN_BORDER

    def _apply_body_style(self, sheet, row: int, columns: int) -> None:
        for column in range(1, columns + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = self._THIN_BORDER
            cell.alignment = self._LEFT if column in {1, 2, 3} else self._CENTER

    def _set_widths(self, sheet, widths: dict[int, int]) -> None:
        for column, width in widths.items():
            sheet.column_dimensions[get_column_letter(column)].width = width

    def _pct_to_decimal(self, value: Any) -> float | None:
        if value is None:
            return None
        return float(value) / 100.0
