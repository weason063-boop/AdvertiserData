import pandas as pd
from openpyxl import load_workbook

from api.database import upsert_billing_history, upsert_client_detail_stats_batch, upsert_client_stats_batch
from api.models import BillingHistory, Client, ClientMonthlyDetailStats, ClientMonthlyStats
from api.services.calculation_service import CalculationService
from api.services.dashboard_service import DashboardService


class TestDashboardService:
    def test_get_latest_month_clients_returns_latest_month_sorted_by_consumption(self, db_session):
        upsert_client_stats_batch(
            "2026-01",
            [
                {"name": "Alpha", "consumption": 90, "fee": 9},
                {"name": "Beta", "consumption": 110, "fee": 11},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-02",
            [
                {"name": "Alpha", "consumption": 150, "fee": 15},
                {"name": "Gamma", "consumption": 80, "fee": 8},
                {"name": "Beta", "consumption": 120, "fee": 12},
            ],
            db=db_session,
        )

        result = DashboardService().get_latest_month_clients(db=db_session)

        assert result["latest_month"] == "2026-02"
        assert [item["client_name"] for item in result["rows"]] == ["Alpha", "Beta", "Gamma"]
        assert result["rows"][0]["consumption"] == 150.0
        assert result["rows"][1]["service_fee"] == 12.0
        assert result["rows"][0]["month"] == "2026-02"
        assert result["rows"][0]["entity"] is None
        assert result["rows"][0]["owner"] is None
        assert result["rows"][0]["bill_amount"] == 165.0
        assert result["rows"][0]["note"] is None

    def test_get_client_history_returns_profile_and_desc_rows(self, db_session):
        db_session.add(
            Client(
                name="Alpha",
                business_type="Agency",
                department="North",
                entity="Entity-A",
                fee_clause="5%",
            )
        )
        db_session.commit()

        upsert_client_stats_batch(
            "2025-12",
            [{"name": "Alpha", "consumption": 100, "fee": 10}],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-01",
            [{"name": "Alpha", "consumption": 110, "fee": 11}],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-02",
            [{"name": "Alpha", "consumption": 120, "fee": 12}],
            db=db_session,
        )

        result = DashboardService().get_client_history("Alpha", db=db_session)

        assert result["profile"]["client_name"] == "Alpha"
        assert result["profile"]["business_type"] == "Agency"
        assert [item["month"] for item in result["rows"]] == ["2026-02", "2026-01", "2025-12"]
        assert result["summary"]["latest_month"] == "2026-02"
        assert result["summary"]["first_month"] == "2025-12"
        assert result["summary"]["total_consumption"] == 330.0
        assert result["summary"]["total_service_fee"] == 33.0

    def test_get_client_history_allows_missing_profile(self, db_session):
        upsert_client_stats_batch(
            "2026-02",
            [{"name": "Missing Profile", "consumption": 88, "fee": 8}],
            db=db_session,
        )

        result = DashboardService().get_client_history("Missing Profile", db=db_session)

        assert result["profile"]["client_name"] == "Missing Profile"
        assert result["profile"]["business_type"] is None
        assert result["summary"]["latest_month"] == "2026-02"
        assert result["rows"][0]["consumption"] == 88.0

    def test_get_latest_month_clients_returns_detail_metrics_when_available(self, db_session):
        db_session.add(
            Client(
                name="Alpha",
                business_type="Agency",
                department="Nora",
                entity="Entity-A",
                fee_clause="5%",
            )
        )
        db_session.commit()

        upsert_client_detail_stats_batch(
            "2026-02",
            [
                {
                    "name": "Alpha",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 90,
                    "managed_consumption": 60,
                    "net_consumption": 145,
                    "service_fee": 8,
                    "fixed_service_fee": 2,
                    "coupon": -5,
                    "dst": 3,
                    "total": 155,
                },
                {
                    "name": "Beta",
                    "bill_type": "后付",
                    "service_type": "直客",
                    "flow_consumption": 40,
                    "managed_consumption": 20,
                    "net_consumption": 58,
                    "service_fee": 3,
                    "fixed_service_fee": 1,
                    "coupon": 0,
                    "dst": 1,
                    "total": 62,
                },
            ],
            db=db_session,
        )

        result = DashboardService().get_latest_month_clients(db=db_session)

        assert result["latest_month"] == "2026-02"
        assert [item["client_name"] for item in result["rows"]] == ["Alpha", "Beta"]
        assert result["rows"][0]["bill_type"] == "预付"
        assert result["rows"][0]["service_type"] == "代投"
        assert result["rows"][0]["consumption"] == 150.0
        assert result["rows"][0]["service_fee_total"] == 10.0
        assert result["rows"][0]["coupon"] == -5.0
        assert result["rows"][0]["month"] == "2026-02"
        assert result["rows"][0]["entity"] == "Entity-A"
        assert result["rows"][0]["owner"] == "Nora"
        assert result["rows"][0]["bill_amount"] == 155.0
        assert result["rows"][0]["note"] is None

    def test_get_latest_month_clients_merges_detail_and_legacy_rows(self, db_session):
        upsert_client_stats_batch(
            "2026-02",
            [
                {"name": "Alpha", "consumption": 120, "fee": 12},
                {"name": "Beta", "consumption": 100, "fee": 10},
            ],
            db=db_session,
        )
        upsert_client_detail_stats_batch(
            "2026-02",
            [
                {
                    "name": "Alpha",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 90,
                    "managed_consumption": 30,
                    "net_consumption": 118,
                    "service_fee": 6,
                    "fixed_service_fee": 2,
                    "coupon": 0,
                    "dst": 1,
                    "total": 128,
                }
            ],
            db=db_session,
        )

        result = DashboardService().get_latest_month_clients(db=db_session)

        assert result["latest_month"] == "2026-02"
        assert [item["client_name"] for item in result["rows"]] == ["Alpha", "Beta"]
        alpha = result["rows"][0]
        beta = result["rows"][1]
        assert alpha["bill_type"] == "预付"
        assert beta["bill_type"] == "—"
        assert beta["consumption"] == 100.0
        assert beta["bill_amount"] == 110.0

    def test_update_client_month_note_persists_and_reflects_in_latest_month(self, db_session):
        upsert_client_stats_batch(
            "2026-02",
            [{"name": "Alpha", "consumption": 120, "fee": 12}],
            db=db_session,
        )
        service = DashboardService()
        payload = service.update_client_month_note(
            month="2026-02",
            client_name="Alpha",
            note="需要复核",
            db=db_session,
        )

        assert payload["month"] == "2026-02"
        assert payload["client_name"] == "Alpha"
        assert payload["note"] == "需要复核"

        latest = service.get_latest_month_clients(db=db_session)
        assert latest["rows"][0]["note"] == "需要复核"

    def test_get_client_history_prefers_detail_rows_and_builds_detail_summary(self, db_session):
        db_session.add(
            Client(
                name="Alpha",
                business_type="Agency",
                department="North",
                entity="Entity-A",
                fee_clause="5%",
            )
        )
        db_session.commit()

        upsert_client_detail_stats_batch(
            "2026-02",
            [
                {
                    "name": "Alpha",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 80,
                    "managed_consumption": 20,
                    "net_consumption": 96,
                    "service_fee": 5,
                    "fixed_service_fee": 1,
                    "coupon": -2,
                    "dst": 3,
                    "total": 103,
                }
            ],
            db=db_session,
        )
        upsert_client_detail_stats_batch(
            "2026-01",
            [
                {
                    "name": "Alpha",
                    "bill_type": "后付",
                    "service_type": "直客",
                    "flow_consumption": 40,
                    "managed_consumption": 10,
                    "net_consumption": 49,
                    "service_fee": 3,
                    "fixed_service_fee": 2,
                    "coupon": 0,
                    "dst": 1,
                    "total": 55,
                }
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2025-12",
            [{"name": "Alpha", "consumption": 90, "fee": 9}],
            db=db_session,
        )

        result = DashboardService().get_client_history("Alpha", db=db_session)

        assert [item["month"] for item in result["rows"]] == ["2026-02", "2026-01", "2025-12"]
        assert result["rows"][0]["bill_type"] == "预付"
        assert result["rows"][1]["service_type"] == "直客"
        assert result["rows"][2]["consumption"] == 90.0
        assert result["summary"]["total_consumption"] == 240.0
        assert result["summary"]["total_service_fee"] == 20.0
        assert result["summary"]["total_total"] == 257.0

    def test_prepare_result_month_batches_aggregates_multi_bill_type_and_dst_variants(self, db_session):
        service = CalculationService()
        source = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 60,
                    "代投消耗": 20,
                    "汇总纯花费": 78,
                    "服务费": 5,
                    "固定服务费": 1,
                    "Coupon": -1,
                    "监管运营费用/数字服务税 (DST)\xa0": 1.5,
                    "汇总": 85.5,
                    "月份归属": "2026-02-01",
                },
                {
                    "母公司": "Alpha",
                    "预付/后付": "后付",
                    "服务类型": "代投",
                    "流水消耗": 40,
                    "代投消耗": 30,
                    "汇总纯花费": 68,
                    "服务费": 4,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "监管运营费用/数字服务税 (DST)\xa0": 1.5,
                    "汇总": 73.5,
                    "月份归属": "2026-02-18",
                },
            ]
        )

        batches = service._prepare_result_month_batches("2026年2月测试_results.xlsx", source)

        assert [month for month, _ in batches] == ["2026-02"]

        month, batch_df = batches[0]
        service._upsert_monthly_stats(month, batch_df, db=db_session)

        result = DashboardService().get_latest_month_clients(db=db_session)
        alpha = next(item for item in result["rows"] if item["client_name"] == "Alpha")

        assert alpha["bill_type"] == "多类型"
        assert alpha["service_type"] == "代投"
        assert alpha["flow_consumption"] == 100.0
        assert alpha["managed_consumption"] == 50.0
        assert alpha["net_consumption"] == 150.0
        assert alpha["dst"] == 3.0
        assert alpha["coupon"] == -1.0
        assert alpha["total"] == 162.0

    def test_prepare_result_month_batches_total_uses_fee_engine_formula(self, db_session):
        service = CalculationService()
        source = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 0,
                    "汇总纯消耗": 100,
                    "换汇汇率": 0.5,
                    "服务费": 1.115,
                    "固定服务费": 0.005,
                    "COUPON": -0.12,
                    "监管运营费用/数字服务税(DST)": 0.3,
                    "月份归属": "2026-02-01",
                }
            ]
        )

        batches = service._prepare_result_month_batches("2026年2月测试_results.xlsx", source)
        month, batch_df = batches[0]
        service._upsert_monthly_stats(month, batch_df, db=db_session)

        result = DashboardService().get_latest_month_clients(db=db_session)
        alpha = next(item for item in result["rows"] if item["client_name"] == "Alpha")

        assert alpha["flow_consumption"] == 0.0
        assert alpha["managed_consumption"] == 0.0
        assert alpha["net_consumption"] == 50.0
        assert alpha["service_fee"] == 1.115
        assert alpha["fixed_service_fee"] == 0.005
        assert alpha["coupon"] == -0.12
        assert alpha["dst"] == 0.3
        assert alpha["total"] == 51.3

    def test_prepare_result_month_batches_total_includes_dst_keyword_variant_header(self, db_session):
        service = CalculationService()
        source = pd.DataFrame(
            [
                {
                    "母公司 ": "Alpha",
                    "预付/后付 ": "预付",
                    "服务类型": "代投",
                    "流水消耗": 80,
                    "代投消耗": 20,
                    "服务费 ": 2,
                    "固定服务费": 1,
                    "coupon": -0.5,
                    "监管费（DST）": 3.2,
                    "月份归属 ": "2026-02-01",
                }
            ]
        )

        batches = service._prepare_result_month_batches("2026年2月测试_results.xlsx", source)
        month, batch_df = batches[0]
        service._upsert_monthly_stats(month, batch_df, db=db_session)

        result = DashboardService().get_latest_month_clients(db=db_session)
        alpha = next(item for item in result["rows"] if item["client_name"] == "Alpha")

        assert alpha["flow_consumption"] == 80.0
        assert alpha["managed_consumption"] == 20.0
        assert alpha["dst"] == 3.2
        assert alpha["coupon"] == -0.5
        assert alpha["total"] == 105.7

    def test_prepare_result_month_batches_prefers_hidden_calc_data_sheet(self, tmp_path):
        service = CalculationService()
        source = tmp_path / "2026年2月测试_results.xlsx"

        visible_df = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "媒介": "Google",
                    "币种": "RMB",
                    "2026年2月消耗": 7200,
                    "2026年2月服务费": 110,
                }
            ]
        )
        hidden_df = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 1100,
                    "汇总纯花费": 1100,
                    "换汇汇率": 1.1 / 7.2,
                    "服务费": 110,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "月份归属": "2026-02-01",
                }
            ]
        )

        with pd.ExcelWriter(source) as writer:
            visible_df.to_excel(writer, index=False, sheet_name="客户端口代投2022.9-2026.3")
            hidden_df.to_excel(writer, index=False, sheet_name=service._RESULT_DATA_SHEET_NAME)

        workbook = load_workbook(source)
        workbook[service._RESULT_DATA_SHEET_NAME].sheet_state = "hidden"
        workbook.save(source)
        workbook.close()

        batches = service._prepare_result_month_batches(source.name, source)

        assert [month for month, _ in batches] == ["2026-02"]
        _, batch_df = batches[0]
        assert batch_df["_managed_consumption"].sum() == 1100.0
        assert batch_df["_service_fee"].sum() == 110.0

    def test_parse_month_from_stored_filename_prefers_original_file_month(self):
        service = CalculationService()

        month = service._parse_month_from_filename(
            "weason_20260424100458247492_44cdd405_2026年3月_results.xlsx"
        )

        assert month == "2026-03"

    def test_parse_month_from_stored_filename_prefers_original_file_month(self):
        service = CalculationService()

        month = service._parse_month_from_filename(
            "weason_20260424100458247492_44cdd405_2026年3月_results.xlsx",
            prefer_latest_match=True,
        )

        assert month == "2026-03"

    def test_parse_month_from_original_filename_prefers_business_month_over_export_date(self):
        service = CalculationService()

        month = service._parse_month_from_filename("2026年3月广告消耗数据_20260425.xlsx")

        assert month == "2026-03"

    def test_prepare_result_month_batches_falls_back_to_all_visible_sheets(self, tmp_path):
        service = CalculationService()
        source = tmp_path / "2026年3月测试_results.xlsx"

        usd_df = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "媒介": "Google",
                    "服务类型": "代投",
                    "代投消耗": 100,
                    "流水消耗": 0,
                    "服务费": 10,
                    "固定服务费": 0,
                    "月份归属": "2026-03-01",
                    "币种": "USD",
                }
            ]
        )
        rmb_df = pd.DataFrame(
            [
                {
                    "母公司": "Beta",
                    "媒介": "LinkedIn",
                    "服务类型": "代投",
                    "代投消耗": 1100,
                    "流水消耗": 0,
                    "服务费": 88,
                    "固定服务费": 0,
                    "月份归属": "2026-03-01",
                    "币种": "RMB",
                }
            ]
        )
        client_account_df = pd.DataFrame(
            [
                {
                    "母公司": "Gamma",
                    "媒介": "Facebook",
                    "币种": "USD",
                    "2026年3月消耗": 500,
                    "2026年3月服务费": 50,
                    "2026年3月固定服务费": 0,
                }
            ]
        )

        with pd.ExcelWriter(source) as writer:
            usd_df.to_excel(writer, index=False, sheet_name="2026年3月美金")
            rmb_df.to_excel(writer, index=False, sheet_name="2026年人民币汇总")
            client_account_df.to_excel(writer, index=False, sheet_name="客户端口账户代投2022.9-2026.3")

        batches = service._prepare_result_month_batches(source.name, source)

        assert [month for month, _ in batches] == ["2026-03"]
        _, batch_df = batches[0]
        assert batch_df["_managed_consumption"].sum() == 1700.0
        assert batch_df["_service_fee"].sum() == 148.0

    def test_prepare_result_month_batches_multi_sheet_uses_filename_month_snapshot(self, tmp_path):
        service = CalculationService()
        source = tmp_path / "2026年3月测试_results.xlsx"

        hidden_df = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "媒介": "Google",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 100,
                    "汇总纯花费": 100,
                    "服务费": 10,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年3月美金",
                    "月份归属": None,
                },
                {
                    "母公司": "Beta",
                    "媒介": "LinkedIn",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 50,
                    "汇总纯花费": 50,
                    "服务费": 5,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年人民币汇总",
                    "月份归属": "2026-03-01",
                },
                {
                    "母公司": "Gamma",
                    "媒介": "LinkedIn",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 30,
                    "汇总纯花费": 30,
                    "服务费": 3,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年人民币汇总",
                    "月份归属": "2026-02-01",
                },
            ]
        )

        with pd.ExcelWriter(source) as writer:
            pd.DataFrame([{"母公司": "Visible", "媒介": "Google"}]).to_excel(
                writer,
                index=False,
                sheet_name="2026年3月美金",
            )
            hidden_df.to_excel(writer, index=False, sheet_name=service._RESULT_DATA_SHEET_NAME)

        workbook = load_workbook(source)
        workbook[service._RESULT_DATA_SHEET_NAME].sheet_state = "hidden"
        workbook.save(source)
        workbook.close()

        batches = service._prepare_result_month_batches(source.name, source)

        assert [month for month, _ in batches] == ["2026-03"]
        _, batch_df = batches[0]
        assert batch_df["_managed_consumption"].sum() == 150.0
        assert batch_df["_service_fee"].sum() == 15.0
        assert sorted(batch_df["_client_name"].tolist()) == ["Alpha", "Beta"]

    def test_upsert_monthly_stats_replaces_stale_month_snapshot(self, db_session):
        service = CalculationService()
        month = "2026-03"

        first_batch = pd.DataFrame(
            [
                {
                    "_client_name": "Alpha",
                    "_bill_type": "预付",
                    "_service_type": "代投",
                    "_flow_consumption": 0.0,
                    "_managed_consumption": 100.0,
                    "_net_consumption": 100.0,
                    "_service_fee": 10.0,
                    "_fixed_service_fee": 0.0,
                    "_coupon": 0.0,
                    "_dst": 0.0,
                    "_total": 110.0,
                    "_temp_consumption": 100.0,
                    "_temp_fee": 10.0,
                },
                {
                    "_client_name": "Beta",
                    "_bill_type": "预付",
                    "_service_type": "代投",
                    "_flow_consumption": 0.0,
                    "_managed_consumption": 200.0,
                    "_net_consumption": 200.0,
                    "_service_fee": 20.0,
                    "_fixed_service_fee": 0.0,
                    "_coupon": 0.0,
                    "_dst": 0.0,
                    "_total": 220.0,
                    "_temp_consumption": 200.0,
                    "_temp_fee": 20.0,
                },
            ]
        )
        service._upsert_monthly_stats(month, first_batch, db=db_session)

        second_batch = pd.DataFrame(
            [
                {
                    "_client_name": "Alpha",
                    "_bill_type": "预付",
                    "_service_type": "代投",
                    "_flow_consumption": 0.0,
                    "_managed_consumption": 50.0,
                    "_net_consumption": 50.0,
                    "_service_fee": 5.0,
                    "_fixed_service_fee": 0.0,
                    "_coupon": 0.0,
                    "_dst": 0.0,
                    "_total": 55.0,
                    "_temp_consumption": 50.0,
                    "_temp_fee": 5.0,
                }
            ]
        )
        service._upsert_monthly_stats(month, second_batch, db=db_session)

        stats_rows = (
            db_session.query(ClientMonthlyStats)
            .filter(ClientMonthlyStats.month == month)
            .order_by(ClientMonthlyStats.client_name)
            .all()
        )
        detail_rows = (
            db_session.query(ClientMonthlyDetailStats)
            .filter(ClientMonthlyDetailStats.month == month)
            .order_by(ClientMonthlyDetailStats.client_name)
            .all()
        )

        assert [row.client_name for row in stats_rows] == ["Alpha"]
        assert [row.client_name for row in detail_rows] == ["Alpha"]
        assert stats_rows[0].consumption == 50.0
        assert stats_rows[0].service_fee == 5.0

        latest = DashboardService().get_latest_month_clients(db=db_session)
        assert [item["client_name"] for item in latest["rows"]] == ["Alpha"]
        assert latest["rows"][0]["consumption"] == 50.0
        assert latest["rows"][0]["service_fee"] == 5.0

    def test_upsert_monthly_stats_omits_zero_only_detail_rows(self, db_session):
        service = CalculationService()
        month = "2026-03"
        batch_df = pd.DataFrame(
            [
                {
                    "_client_name": "Active",
                    "_bill_type": "预付",
                    "_service_type": "代投",
                    "_flow_consumption": 0.0,
                    "_managed_consumption": 100.0,
                    "_net_consumption": 100.0,
                    "_service_fee": 10.0,
                    "_fixed_service_fee": 0.0,
                    "_coupon": 0.0,
                    "_dst": 0.0,
                    "_total": 110.0,
                    "_temp_consumption": 100.0,
                    "_temp_fee": 10.0,
                },
                {
                    "_client_name": "Zero",
                    "_bill_type": "预付",
                    "_service_type": "代投",
                    "_flow_consumption": 0.0,
                    "_managed_consumption": 0.0,
                    "_net_consumption": 0.0,
                    "_service_fee": 0.0,
                    "_fixed_service_fee": 0.0,
                    "_coupon": 0.0,
                    "_dst": 0.0,
                    "_total": 0.0,
                    "_temp_consumption": 0.0,
                    "_temp_fee": 0.0,
                },
            ]
        )

        service._upsert_monthly_stats(month, batch_df, db=db_session)

        detail_rows = (
            db_session.query(ClientMonthlyDetailStats)
            .filter(ClientMonthlyDetailStats.month == month)
            .order_by(ClientMonthlyDetailStats.client_name)
            .all()
        )

        assert [row.client_name for row in detail_rows] == ["Active"]

    def test_upsert_monthly_stats_removes_zero_only_billing_history_rows(self, db_session):
        service = CalculationService()
        month = "2026-04"
        upsert_billing_history(month, 999, 99, db=db_session)

        batch_df = pd.DataFrame(
            [
                {
                    "_client_name": "Zero",
                    "_bill_type": "æ£°å‹ªç²¯",
                    "_service_type": "æµ ï½†å§‡",
                    "_flow_consumption": 0.0,
                    "_managed_consumption": 0.0,
                    "_net_consumption": 0.0,
                    "_service_fee": 0.0,
                    "_fixed_service_fee": 0.0,
                    "_coupon": 0.0,
                    "_dst": 0.0,
                    "_total": 0.0,
                    "_temp_consumption": 0.0,
                    "_temp_fee": 0.0,
                }
            ]
        )

        service._upsert_monthly_stats(month, batch_df, db=db_session)

        billing_row = (
            db_session.query(BillingHistory)
            .filter(BillingHistory.month == month)
            .first()
        )

        assert billing_row is None

    def test_backfill_detail_stats_from_results_rebuilds_month_snapshots(self, db_session, monkeypatch, tmp_path):
        service = CalculationService()
        source = tmp_path / "2026年3月测试_results.xlsx"

        hidden_df = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "媒介": "Google",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 100,
                    "汇总纯花费": 100,
                    "服务费": 10,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年3月美金",
                    "月份归属": None,
                },
                {
                    "母公司": "Beta",
                    "媒介": "LinkedIn",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 50,
                    "汇总纯花费": 50,
                    "服务费": 5,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年人民币汇总",
                    "月份归属": "2026-03-01",
                },
                {
                    "母公司": "Gamma",
                    "媒介": "LinkedIn",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 30,
                    "汇总纯花费": 30,
                    "服务费": 3,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年人民币汇总",
                    "月份归属": "2026-02-01",
                },
            ]
        )

        with pd.ExcelWriter(source) as writer:
            pd.DataFrame([{"母公司": "Visible", "媒介": "Google"}]).to_excel(
                writer,
                index=False,
                sheet_name="2026年3月美金",
            )
            hidden_df.to_excel(writer, index=False, sheet_name=service._RESULT_DATA_SHEET_NAME)

        workbook = load_workbook(source)
        workbook[service._RESULT_DATA_SHEET_NAME].sheet_state = "hidden"
        workbook.save(source)
        workbook.close()

        upsert_client_stats_batch("2026-01", [{"name": "Stale", "consumption": 999, "fee": 99}], db=db_session)
        upsert_client_detail_stats_batch(
            "2026-01",
            [
                {
                    "name": "Stale",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 0,
                    "managed_consumption": 999,
                    "net_consumption": 999,
                    "service_fee": 99,
                    "fixed_service_fee": 0,
                    "coupon": 0,
                    "dst": 0,
                    "total": 1098,
                }
            ],
            db=db_session,
        )
        upsert_billing_history("2026-01", 999, 99, db=db_session)

        monkeypatch.setenv("TESTING", "False")
        monkeypatch.setattr(service, "_get_upload_dir", lambda: tmp_path)

        service.backfill_detail_stats_from_results(db=db_session)

        stats_rows = (
            db_session.query(ClientMonthlyStats)
            .order_by(ClientMonthlyStats.month, ClientMonthlyStats.client_name)
            .all()
        )
        detail_rows = (
            db_session.query(ClientMonthlyDetailStats)
            .order_by(ClientMonthlyDetailStats.month, ClientMonthlyDetailStats.client_name)
            .all()
        )

        assert {(row.month, row.client_name) for row in stats_rows} == {
            ("2026-03", "Alpha"),
            ("2026-03", "Beta"),
        }
        assert {(row.month, row.client_name) for row in detail_rows} == {
            ("2026-03", "Alpha"),
            ("2026-03", "Beta"),
        }

        billing_rows = (
            db_session.query(BillingHistory)
            .order_by(BillingHistory.month.asc())
            .all()
        )
        assert [(row.month, row.total_consumption, row.total_service_fee) for row in billing_rows] == [
            ("2026-03", 150.0, 15.0),
        ]

    def test_backfill_detail_stats_from_results_uses_registered_source_file_and_ignores_estimates(
        self,
        db_session,
        monkeypatch,
        tmp_path,
    ):
        service = CalculationService()
        source = tmp_path / "system_20260424100458247492_44cdd405_2026年3月_results.xlsx"
        estimate = tmp_path / "system_20260424100458247493_44cdd405_invalid_estimate_results.xlsx"

        hidden_df = pd.DataFrame(
            [
                {
                    "母公司": "Alpha",
                    "媒介": "Google",
                    "预付/后付": "预付",
                    "服务类型": "代投",
                    "流水消耗": 0,
                    "代投消耗": 100,
                    "汇总纯花费": 100,
                    "服务费": 10,
                    "固定服务费": 0,
                    "Coupon": 0,
                    "来源Sheet": "2026年3月美金",
                    "月份归属": None,
                }
            ]
        )

        with pd.ExcelWriter(source) as writer:
            pd.DataFrame([{"母公司": "Visible", "媒介": "Google"}]).to_excel(
                writer,
                index=False,
                sheet_name="2026年3月美金",
            )
            hidden_df.to_excel(writer, index=False, sheet_name=service._RESULT_DATA_SHEET_NAME)

        workbook = load_workbook(source)
        workbook[service._RESULT_DATA_SHEET_NAME].sheet_state = "hidden"
        workbook.save(source)
        workbook.close()

        with pd.ExcelWriter(estimate) as writer:
            pd.DataFrame([{"foo": 1}]).to_excel(writer, index=False, sheet_name="Sheet1")

        monkeypatch.setenv("TESTING", "False")
        monkeypatch.setattr(service, "_get_upload_dir", lambda: tmp_path)
        service._register_result(
            filename=source.name,
            owner_username="system",
            source_file="2026年3月广告消耗数据_20260425.xlsx",
            operation="calculate",
        )
        service._register_result(
            filename=estimate.name,
            owner_username="system",
            source_file="4月预估测试.xlsx",
            operation="estimate_calculate",
        )

        service.backfill_detail_stats_from_results(db=db_session)

        billing_rows = (
            db_session.query(BillingHistory)
            .order_by(BillingHistory.month.asc())
            .all()
        )
        assert [(row.month, row.total_consumption, row.total_service_fee) for row in billing_rows] == [
            ("2026-03", 100.0, 10.0),
        ]

    def test_backfill_detail_stats_from_results_preserves_existing_snapshot_when_any_result_fails(
        self,
        db_session,
        monkeypatch,
        tmp_path,
    ):
        service = CalculationService()
        bad_result = tmp_path / "system_20260424100458247492_44cdd405_bad_results.xlsx"

        with pd.ExcelWriter(bad_result) as writer:
            pd.DataFrame([{"母公司": "Broken", "媒介": "Google"}]).to_excel(
                writer,
                index=False,
                sheet_name="Sheet1",
            )

        upsert_client_stats_batch("2026-01", [{"name": "Stale", "consumption": 999, "fee": 99}], db=db_session)
        upsert_client_detail_stats_batch(
            "2026-01",
            [
                {
                    "name": "Stale",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 0,
                    "managed_consumption": 999,
                    "net_consumption": 999,
                    "service_fee": 99,
                    "fixed_service_fee": 0,
                    "coupon": 0,
                    "dst": 0,
                    "total": 1098,
                }
            ],
            db=db_session,
        )
        upsert_billing_history("2026-01", 999, 99, db=db_session)

        monkeypatch.setenv("TESTING", "False")
        monkeypatch.setattr(service, "_get_upload_dir", lambda: tmp_path)
        service._register_result(
            filename=bad_result.name,
            owner_username="system",
            source_file="工作簿1.xlsx",
            operation="calculate",
        )

        service.backfill_detail_stats_from_results(db=db_session)

        billing_rows = (
            db_session.query(BillingHistory)
            .order_by(BillingHistory.month.asc())
            .all()
        )
        assert [(row.month, row.total_consumption, row.total_service_fee) for row in billing_rows] == [
            ("2026-01", 999.0, 99.0),
        ]

    def test_get_main_stats_triggers_result_backfill_sync(self, db_session, monkeypatch):
        class DummyCalculationService:
            def __init__(self):
                self.calls = 0

            def backfill_detail_stats_from_results(self, db=None):
                self.calls += 1

        calc_service = DummyCalculationService()
        service = DashboardService(calculation_service=calc_service)
        monkeypatch.setenv("TESTING", "False")

        upsert_billing_history("2026-03", 1000, 100, db=db_session)

        payload = service.get_main_stats(db=db_session)

        assert calc_service.calls == 1
        assert payload["stats"]["month"] == "2026-03"
        assert payload["stats"]["consumption"] == 1000
        assert payload["stats"]["fee"] == 100

    def test_get_main_stats_ignores_zero_only_latest_month_in_trend(self, db_session, monkeypatch):
        class DummyCalculationService:
            def backfill_detail_stats_from_results(self, db=None):
                return None

        service = DashboardService(calculation_service=DummyCalculationService())
        monkeypatch.setenv("TESTING", "False")

        upsert_billing_history("2026-02", 800, 80, db=db_session)
        upsert_billing_history("2026-03", 1000, 100, db=db_session)
        upsert_billing_history("2026-04", 0, 0, db=db_session)

        payload = service.get_main_stats(db=db_session)

        assert payload["stats"]["month"] == "2026-03"
        assert payload["stats"]["consumption"] == 1000
        assert payload["stats"]["fee"] == 100
        assert [item["month"] for item in payload["trend"]] == ["2026-02", "2026-03"]

    def test_get_latest_month_clients_prefers_latest_active_billing_month(self, db_session, monkeypatch):
        class DummyCalculationService:
            def backfill_detail_stats_from_results(self, db=None):
                return None

        service = DashboardService(calculation_service=DummyCalculationService())
        monkeypatch.setenv("TESTING", "False")

        upsert_client_stats_batch(
            "2026-03",
            [{"name": "Alpha", "consumption": 100, "fee": 10}],
            db=db_session,
        )
        upsert_client_detail_stats_batch(
            "2026-03",
            [
                {
                    "name": "Alpha",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 0,
                    "managed_consumption": 100,
                    "net_consumption": 100,
                    "service_fee": 10,
                    "fixed_service_fee": 0,
                    "coupon": 0,
                    "dst": 0,
                    "total": 110,
                }
            ],
            db=db_session,
        )
        upsert_billing_history("2026-03", 100, 10, db=db_session)
        upsert_client_detail_stats_batch(
            "2026-04",
            [
                {
                    "name": "Alpha",
                    "bill_type": "预付",
                    "service_type": "代投",
                    "flow_consumption": 0,
                    "managed_consumption": 0,
                    "net_consumption": 0,
                    "service_fee": 0,
                    "fixed_service_fee": 0,
                    "coupon": -5,
                    "dst": 0,
                    "total": -5,
                }
            ],
            db=db_session,
        )

        payload = service.get_latest_month_clients(db=db_session)

        assert payload["latest_month"] == "2026-03"
        assert [item["client_name"] for item in payload["rows"]] == ["Alpha"]

    def test_client_trend_summary_uses_visible_12_month_window(self, db_session):
        upsert_client_stats_batch(
            "1970-01",
            [{"name": "Test Client", "consumption": 999999, "fee": 0}],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2025-12",
            [{"name": "Test Client", "consumption": 200, "fee": 20}],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-01",
            [{"name": "Test Client", "consumption": 500, "fee": 50}],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-02",
            [{"name": "Test Client", "consumption": 300, "fee": 30}],
            db=db_session,
        )

        result = DashboardService().get_client_trend("Test Client", db=db_session)

        assert len(result["data"]) == 12
        assert result["summary"]["total_consumption"] == 1000
        assert result["summary"]["avg_monthly"] == 333.33
        assert result["summary"]["peak_month"] == "2026-01"
        assert result["summary"]["peak_value"] == 500

    def test_quarter_top_clients_supports_dual_compare_mode(self, db_session):
        upsert_client_stats_batch(
            "2025-03",
            [
                {"name": "Alpha", "consumption": 50, "fee": 5},
                {"name": "Beta", "consumption": 30, "fee": 3},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2025-04",
            [
                {"name": "Alpha", "consumption": 60, "fee": 6},
                {"name": "Beta", "consumption": 20, "fee": 2},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2025-05",
            [
                {"name": "Alpha", "consumption": 70, "fee": 7},
                {"name": "Beta", "consumption": 10, "fee": 1},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2025-12",
            [
                {"name": "Alpha", "consumption": 40, "fee": 4},
                {"name": "Beta", "consumption": 50, "fee": 5},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-01",
            [
                {"name": "Alpha", "consumption": 40, "fee": 4},
                {"name": "Beta", "consumption": 50, "fee": 5},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-02",
            [
                {"name": "Alpha", "consumption": 40, "fee": 4},
                {"name": "Beta", "consumption": 50, "fee": 5},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-03",
            [
                {"name": "Alpha", "consumption": 100, "fee": 10},
                {"name": "Beta", "consumption": 80, "fee": 8},
                {"name": "Gamma", "consumption": 20, "fee": 2},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-04",
            [
                {"name": "Alpha", "consumption": 110, "fee": 11},
                {"name": "Beta", "consumption": 70, "fee": 7},
                {"name": "Gamma", "consumption": 30, "fee": 3},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-05",
            [
                {"name": "Alpha", "consumption": 90, "fee": 9},
                {"name": "Beta", "consumption": 90, "fee": 9},
                {"name": "Gamma", "consumption": 40, "fee": 4},
            ],
            db=db_session,
        )

        result = DashboardService().get_quarter_top_clients(
            "2026-Q1",
            10,
            db=db_session,
            compare_mode="dual",
        )

        assert result["compare_mode"] == "dual"
        assert result["prev_quarter"] == "2025-Q4"
        assert result["yoy_quarter"] == "2025-Q1"

        alpha = next(item for item in result["clients"] if item["client_name"] == "Alpha")
        beta = next(item for item in result["clients"] if item["client_name"] == "Beta")

        assert alpha["consumption"] == 300.0
        assert alpha["prev_quarter_consumption"] == 120.0
        assert alpha["yoy_consumption"] == 180.0
        assert alpha["qoq_delta"] == 180.0
        assert alpha["yoy_delta"] == 120.0
        assert alpha["qoq_rank_change"] == 1
        assert alpha["yoy_rank_change"] == 0

        assert beta["consumption"] == 240.0
        assert beta["prev_quarter_consumption"] == 150.0
        assert beta["yoy_consumption"] == 60.0
        assert beta["qoq_rank_change"] == -1
        assert beta["yoy_rank_change"] == 0

    def test_month_top_clients_supports_dual_compare_mode(self, db_session):
        upsert_client_stats_batch(
            "2025-02",
            [
                {"name": "Alpha", "consumption": 120, "fee": 12},
                {"name": "Beta", "consumption": 20, "fee": 2},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-01",
            [
                {"name": "Alpha", "consumption": 90, "fee": 9},
                {"name": "Beta", "consumption": 110, "fee": 11},
                {"name": "Gamma", "consumption": 40, "fee": 4},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-02",
            [
                {"name": "Alpha", "consumption": 150, "fee": 15},
                {"name": "Beta", "consumption": 80, "fee": 8},
                {"name": "Gamma", "consumption": 60, "fee": 6},
            ],
            db=db_session,
        )

        result = DashboardService().get_month_top_clients(
            "2026-02",
            10,
            db=db_session,
            compare_mode="dual",
        )

        assert result["compare_mode"] == "dual"
        assert result["prev_month"] == "2026-01"
        assert result["yoy_month"] == "2025-02"

        alpha = next(item for item in result["clients"] if item["client_name"] == "Alpha")
        gamma = next(item for item in result["clients"] if item["client_name"] == "Gamma")

        assert alpha["prev_month_consumption"] == 90.0
        assert alpha["yoy_consumption"] == 120.0
        assert alpha["mom_delta"] == 60.0
        assert alpha["yoy_delta"] == 30.0
        assert alpha["mom_rank_change"] == 1
        assert alpha["yoy_rank_change"] == 0

        assert gamma["prev_month_consumption"] == 40.0
        assert gamma["yoy_consumption"] is None
        assert gamma["mom_delta"] == 20.0
        assert gamma["yoy_delta"] is None
