from io import BytesIO

from openpyxl import load_workbook

from api.database import upsert_billing_history, upsert_client_stats_batch
from api.services.dashboard_report_service import DashboardReportService


class TestDashboardReportService:
    def test_build_month_report_contains_summary_top_clients_and_details(self, db_session):
        upsert_billing_history("2025-03", 800, 80, db=db_session)
        upsert_billing_history("2026-02", 900, 90, db=db_session)
        upsert_billing_history("2026-03", 1000, 100, db=db_session)

        upsert_client_stats_batch(
            "2026-02",
            [
                {"name": "Alpha", "consumption": 500, "fee": 50},
                {"name": "Beta", "consumption": 400, "fee": 40},
            ],
            db=db_session,
        )
        upsert_client_stats_batch(
            "2026-03",
            [
                {"name": "Alpha", "consumption": 600, "fee": 60},
                {"name": "Beta", "consumption": 250, "fee": 25},
                {"name": "Gamma", "consumption": 150, "fee": 15},
            ],
            db=db_session,
        )

        buffer, filename = DashboardReportService().build_report(
            period_type="month",
            period="2026-03",
            limit=2,
            include_details=True,
            db=db_session,
        )

        workbook = load_workbook(BytesIO(buffer.getvalue()), data_only=True)
        assert filename == "dashboard_report_2026-03.xlsx"
        assert workbook.sheetnames == ["Summary", "Top Clients", "Details"]
        assert workbook["Summary"]["A1"].value == "对外汇报看板报表"
        assert workbook["Summary"]["B10"].value == 1000
        assert workbook["Summary"]["B11"].value == 100
        assert workbook["Top Clients"]["B6"].value == "Alpha"
        assert workbook["Top Clients"]["C6"].value == 600
        assert workbook["Details"]["B8"].value == "Gamma"

    def test_build_quarter_report_aggregates_client_details(self, db_session):
        for month, consumption, fee in (
            ("2025-03", 400, 40),
            ("2025-04", 500, 50),
            ("2025-05", 600, 60),
            ("2026-03", 700, 70),
            ("2026-04", 800, 80),
            ("2026-05", 900, 90),
        ):
            upsert_billing_history(month, consumption, fee, db=db_session)

        for month, stats in (
            ("2025-03", [{"name": "Alpha", "consumption": 250, "fee": 25}, {"name": "Beta", "consumption": 150, "fee": 15}]),
            ("2025-04", [{"name": "Alpha", "consumption": 300, "fee": 30}, {"name": "Beta", "consumption": 200, "fee": 20}]),
            ("2025-05", [{"name": "Alpha", "consumption": 350, "fee": 35}, {"name": "Beta", "consumption": 250, "fee": 25}]),
            ("2026-03", [{"name": "Alpha", "consumption": 400, "fee": 40}, {"name": "Beta", "consumption": 300, "fee": 30}]),
            ("2026-04", [{"name": "Alpha", "consumption": 450, "fee": 45}, {"name": "Beta", "consumption": 350, "fee": 35}]),
            ("2026-05", [{"name": "Alpha", "consumption": 500, "fee": 50}, {"name": "Beta", "consumption": 400, "fee": 40}]),
        ):
            upsert_client_stats_batch(month, stats, db=db_session)

        buffer, _ = DashboardReportService().build_report(
            period_type="quarter",
            period="2026-Q1",
            limit=10,
            include_details=True,
            db=db_session,
        )

        workbook = load_workbook(BytesIO(buffer.getvalue()), data_only=True)
        assert workbook["Summary"]["B19"].value == "2026-Q1"
        assert workbook["Summary"]["B20"].value == "2026-03, 2026-04, 2026-05"
        assert workbook["Summary"]["B21"].value == 2400
        assert workbook["Top Clients"]["B6"].value == "Alpha"
        assert workbook["Top Clients"]["C6"].value == 1350
        assert workbook["Top Clients"]["E6"].value == 900
