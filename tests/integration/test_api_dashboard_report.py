from io import BytesIO

from openpyxl import load_workbook

from api.auth import get_current_user_info
from api.main import app
from api.database import upsert_billing_history, upsert_client_stats_batch
import api.routers.dashboard as dashboard_router


BILLING_USER = {
    "id": 1,
    "username": "billing",
    "role": "user",
    "permissions": ["billing_run"],
}
READONLY_USER = {
    "id": 2,
    "username": "readonly",
    "role": "user",
    "permissions": [],
}


def _override_current_user_info(payload: dict):
    app.dependency_overrides[get_current_user_info] = lambda: payload


def _clear_overrides():
    app.dependency_overrides.pop(get_current_user_info, None)


def _seed_dashboard_data(db_session):
    upsert_billing_history("2026-02", 900, 90, db=db_session)
    upsert_billing_history("2026-03", 1000, 100, db=db_session)
    upsert_client_stats_batch(
        "2026-02",
        [
            {"name": "Alpha", "consumption": 500, "fee": 50},
            {"name": "Beta", "consumption": 400, "fee": 40},
        ],
        db=db_session,
    )
    upsert_client_stats_batch(
        "2026-03",
        [
            {"name": "Alpha", "consumption": 600, "fee": 60},
            {"name": "Beta", "consumption": 250, "fee": 25},
            {"name": "Gamma", "consumption": 150, "fee": 15},
        ],
        db=db_session,
    )


def test_dashboard_report_requires_authentication(client, db_session):
    _seed_dashboard_data(db_session)

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-03")

    assert response.status_code == 401


def test_dashboard_report_requires_billing_permission(client, db_session):
    _seed_dashboard_data(db_session)
    _override_current_user_info(READONLY_USER)

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-03")

    assert response.status_code == 403
    _clear_overrides()


def test_dashboard_report_downloads_workbook_for_billing_user(client, db_session):
    _seed_dashboard_data(db_session)
    _override_current_user_info(BILLING_USER)

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-03&limit=2&include_details=true")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert 'dashboard_report_2026-03.xlsx' in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Summary", "Top Clients", "Details"]
    assert workbook["Summary"]["B10"].value == 1000
    assert workbook["Top Clients"]["B6"].value == "Alpha"
    _clear_overrides()


def test_dashboard_report_rejects_invalid_period_format(client, db_session):
    _seed_dashboard_data(db_session)
    _override_current_user_info(BILLING_USER)

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-13")

    assert response.status_code == 400
    _clear_overrides()


def test_dashboard_report_returns_empty_workbook_for_valid_period_without_data(client, db_session):
    _override_current_user_info(BILLING_USER)

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-04")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["Summary", "Top Clients", "Details"]
    assert workbook["Top Clients"]["A6"].value == "暂无数据"
    _clear_overrides()


def test_dashboard_report_records_audit_on_success(client, db_session, monkeypatch):
    _seed_dashboard_data(db_session)
    _override_current_user_info(BILLING_USER)
    captured: list[dict] = []

    monkeypatch.setattr(
        dashboard_router,
        "record_operation_audit",
        lambda **kwargs: captured.append(kwargs),
    )

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-03")

    assert response.status_code == 200
    assert captured
    event = captured[-1]
    assert event["category"] == "dashboard"
    assert event["action"] == "export_report"
    assert event["actor"] == "billing"
    assert event["status"] == "success"
    _clear_overrides()


def test_dashboard_report_records_audit_on_failed_export(client, db_session, monkeypatch):
    _seed_dashboard_data(db_session)
    _override_current_user_info(BILLING_USER)
    captured: list[dict] = []

    monkeypatch.setattr(
        dashboard_router,
        "record_operation_audit",
        lambda **kwargs: captured.append(kwargs),
    )

    response = client.get("/api/dashboard/export/report.xlsx?period_type=month&period=2026-99")

    assert response.status_code == 400
    assert captured
    event = captured[-1]
    assert event["category"] == "dashboard"
    assert event["action"] == "export_report"
    assert event["status"] == "failed"
    _clear_overrides()
